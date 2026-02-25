"""
SAP Security Report Tools for MCP Server
Generates Excel reports and risk summaries from security analysis.

Uses openpyxl for Excel generation with professional formatting.
"""
import base64
import io
import logging
from datetime import datetime
from typing import Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .security_tools import SecurityTools
from .basis_tools import BasisTools

logger = logging.getLogger(__name__)


class ReportTools:
    """Report generation tools for SAP security analysis."""

    def __init__(self, sap_connection):
        """
        Initialize with SAP connection.

        Args:
            sap_connection: SAPRestConnector instance (already connected)
        """
        self.conn = sap_connection
        self.security_tools = SecurityTools(sap_connection)
        self.basis_tools = BasisTools(sap_connection)

    def generate_security_excel(self, report_type: str = "full_report") -> Dict[str, Any]:
        """
        Tool 16: Generate Excel security report.

        Args:
            report_type: Type of report to generate
                - dormant_users: Dormant user analysis
                - sod_violations: Segregation of Duties report
                - critical_access: Critical transaction access report
                - full_report: Complete security assessment

        Returns:
            Dict with base64 encoded Excel file and filename
        """
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl library not installed. Install with: pip install openpyxl"
            }

        try:
            # Create workbook
            wb = Workbook()

            # Generate report based on type
            if report_type == "dormant_users":
                self._generate_dormant_users_report(wb)
                filename = f"SAP_Dormant_Users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            elif report_type == "sod_violations":
                self._generate_sod_report(wb)
                filename = f"SAP_SoD_Violations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            elif report_type == "critical_access":
                self._generate_critical_access_report(wb)
                filename = f"SAP_Critical_Access_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            else:  # full_report
                self._generate_full_report(wb)
                filename = f"SAP_Security_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Encode to base64
            excel_content = base64.b64encode(buffer.getvalue()).decode('utf-8')

            return {
                "success": True,
                "filename": filename,
                "content_base64": excel_content,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "message": f"Excel report generated: {filename}"
            }

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return {"success": False, "error": str(e)}

    def generate_risk_summary(self) -> Dict[str, Any]:
        """
        Tool 17: Generate comprehensive risk summary.

        Returns:
            Dict with aggregated risk summary and recommendations
        """
        try:
            # Collect data from all security tools
            sap_all = self.security_tools.check_sap_all_users()
            dormant = self.security_tools.get_dormant_users(days=90)
            locked = self.security_tools.get_locked_users()
            sod = self.security_tools.check_sod_violations()
            defaults = self.security_tools.check_default_users()
            password_policy = self.security_tools.check_password_policy()
            failed_jobs = self.basis_tools.get_failed_jobs(hours=24)
            system_info = self.basis_tools.get_system_info()

            # Calculate risk metrics
            critical_issues = 0
            high_issues = 0
            medium_issues = 0

            findings = []

            # SAP_ALL users
            if sap_all.get("success") and sap_all.get("users"):
                count = len(sap_all["users"])
                critical_issues += count
                findings.append({
                    "category": "SAP_ALL/SAP_NEW Users",
                    "count": count,
                    "risk": "CRITICAL",
                    "description": f"{count} users have SAP_ALL or SAP_NEW profiles",
                    "action": "Remove SAP_ALL/SAP_NEW profiles immediately"
                })

            # Dormant users
            if dormant.get("success") and dormant.get("users"):
                count = len(dormant["users"])
                if count > 50:
                    high_issues += 1
                    findings.append({
                        "category": "Dormant Users",
                        "count": count,
                        "risk": "HIGH",
                        "description": f"{count} users inactive for 90+ days",
                        "action": "Review and lock inactive accounts"
                    })
                elif count > 0:
                    medium_issues += 1
                    findings.append({
                        "category": "Dormant Users",
                        "count": count,
                        "risk": "MEDIUM",
                        "description": f"{count} users inactive for 90+ days",
                        "action": "Review and lock inactive accounts"
                    })

            # SoD violations
            if sod.get("success") and sod.get("violations"):
                count = len(sod["violations"])
                critical_sod = sum(1 for v in sod["violations"] if v.get("risk") == "CRITICAL")
                if critical_sod > 0:
                    critical_issues += critical_sod
                    findings.append({
                        "category": "SoD Violations (Critical)",
                        "count": critical_sod,
                        "risk": "CRITICAL",
                        "description": f"{critical_sod} critical segregation of duties violations",
                        "action": "Remediate conflicting access immediately"
                    })
                if count > critical_sod:
                    high_issues += (count - critical_sod)
                    findings.append({
                        "category": "SoD Violations (High)",
                        "count": count - critical_sod,
                        "risk": "HIGH",
                        "description": f"{count - critical_sod} high-risk SoD violations",
                        "action": "Review and remediate conflicting roles"
                    })

            # Default users
            if defaults.get("success") and defaults.get("results"):
                critical_defaults = [u for u in defaults["results"]
                                     if u.get("risk") == "CRITICAL"]
                if critical_defaults:
                    critical_issues += len(critical_defaults)
                    findings.append({
                        "category": "Default User Risk",
                        "count": len(critical_defaults),
                        "risk": "CRITICAL",
                        "description": f"{len(critical_defaults)} default users with critical issues",
                        "action": "Lock default users and change passwords"
                    })

            # Password policy
            if password_policy.get("success"):
                non_compliant = sum(1 for p in password_policy.get("parameters", [])
                                    if not p.get("compliant", True))
                if non_compliant > 3:
                    high_issues += 1
                    findings.append({
                        "category": "Password Policy",
                        "count": non_compliant,
                        "risk": "HIGH",
                        "description": f"{non_compliant} password parameters non-compliant",
                        "action": "Update password policy parameters"
                    })
                elif non_compliant > 0:
                    medium_issues += 1
                    findings.append({
                        "category": "Password Policy",
                        "count": non_compliant,
                        "risk": "MEDIUM",
                        "description": f"{non_compliant} password parameters need review",
                        "action": "Review and update password settings"
                    })

            # Failed jobs
            if failed_jobs.get("success") and failed_jobs.get("jobs"):
                count = len(failed_jobs["jobs"])
                if count > 10:
                    medium_issues += 1
                    findings.append({
                        "category": "Failed Background Jobs",
                        "count": count,
                        "risk": "MEDIUM",
                        "description": f"{count} jobs failed in last 24 hours",
                        "action": "Investigate and resolve job failures"
                    })

            # Calculate overall risk score
            if critical_issues > 0:
                overall_risk = "CRITICAL"
                risk_score = 10
            elif high_issues > 3:
                overall_risk = "HIGH"
                risk_score = 8
            elif high_issues > 0 or medium_issues > 3:
                overall_risk = "MEDIUM"
                risk_score = 5
            else:
                overall_risk = "LOW"
                risk_score = 2

            # Get total user counts
            total_locked = len(locked.get("locked_users", [])) if locked.get("success") else 0
            total_dormant = len(dormant.get("users", [])) if dormant.get("success") else 0

            # Top 5 actions
            top_actions = [f["action"] for f in sorted(
                findings,
                key=lambda x: {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2}.get(x["risk"], 3)
            )][:5]

            return {
                "success": True,
                "generated_at": datetime.now().isoformat(),
                "system_info": system_info.get("summary", "Unknown") if system_info.get("success") else "Unknown",
                "overall_risk": overall_risk,
                "risk_score": risk_score,
                "summary": {
                    "critical_issues": critical_issues,
                    "high_issues": high_issues,
                    "medium_issues": medium_issues,
                    "total_locked_users": total_locked,
                    "total_dormant_users": total_dormant
                },
                "findings": findings,
                "top_5_actions": top_actions,
                "recommendation": f"Overall risk level is {overall_risk}. "
                                  f"Address {critical_issues} critical and {high_issues} high-risk issues first."
            }

        except Exception as e:
            logger.error(f"Error generating risk summary: {e}")
            return {"success": False, "error": str(e)}

    # Excel generation helpers

    def _apply_header_style(self, ws, row: int, cols: int):
        """Apply header row styling."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _apply_risk_color(self, ws, cell, risk: str):
        """Apply color based on risk level."""
        colors = {
            "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "HIGH": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "LOW": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        }

        if risk in colors:
            cell.fill = colors[risk]
            if risk in ["CRITICAL", "HIGH"]:
                cell.font = Font(bold=True, color="FFFFFF" if risk == "CRITICAL" else "000000")

    def _auto_column_width(self, ws):
        """Auto-adjust column widths."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _add_report_header(self, ws, title: str):
        """Add report header with title and timestamp."""
        ws.merge_cells('A1:F1')
        ws['A1'] = f"SAP Security Report - {title}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws['A2'] = f"Generated by SyntaAI on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

    def _generate_dormant_users_report(self, wb):
        """Generate dormant users report."""
        ws = wb.active
        ws.title = "Dormant Users"

        self._add_report_header(ws, "Dormant Users Analysis")

        # Get data
        dormant = self.security_tools.get_dormant_users(days=90)

        # Add summary
        ws['A4'] = "Summary"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = f"Total Dormant Users (90+ days): {dormant.get('total_dormant_users', 0)}"
        ws['A6'] = f"Risk Level: {dormant.get('risk_level', 'Unknown')}"

        # Add data table
        headers = ["Username", "User Type", "Last Login", "Days Inactive", "Created On"]
        row = 8

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))

        row += 1
        for user in dormant.get("users", []):
            ws.cell(row=row, column=1, value=user.get("username", ""))
            ws.cell(row=row, column=2, value=user.get("user_type", ""))
            ws.cell(row=row, column=3, value=user.get("last_login", ""))
            ws.cell(row=row, column=4, value=str(user.get("days_inactive", "")))
            ws.cell(row=row, column=5, value=user.get("created_on", ""))
            row += 1

        self._auto_column_width(ws)

    def _generate_sod_report(self, wb):
        """Generate SoD violations report."""
        ws = wb.active
        ws.title = "SoD Violations"

        self._add_report_header(ws, "Segregation of Duties Violations")

        # Get data
        sod = self.security_tools.check_sod_violations()

        # Add summary
        ws['A4'] = "Summary"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = f"Total Violations: {sod.get('total_violations', 0)}"
        ws['A6'] = f"Risk Level: {sod.get('risk_level', 'Unknown')}"
        ws['A7'] = f"Rules Checked: {', '.join(sod.get('rules_checked', []))}"

        # Add data table
        headers = ["User", "Rule", "Risk", "Description", "Conflicting TCodes", "Via Roles"]
        row = 9

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, len(headers))

        row += 1
        for violation in sod.get("violations", []):
            ws.cell(row=row, column=1, value=violation.get("user", ""))
            ws.cell(row=row, column=2, value=violation.get("rule_name", ""))
            risk_cell = ws.cell(row=row, column=3, value=violation.get("risk", ""))
            self._apply_risk_color(ws, risk_cell, violation.get("risk", ""))
            ws.cell(row=row, column=4, value=violation.get("description", ""))
            ws.cell(row=row, column=5, value=", ".join(violation.get("conflicting_tcodes", [])))
            ws.cell(row=row, column=6, value=", ".join(violation.get("via_roles", [])[:3]))
            row += 1

        self._auto_column_width(ws)

    def _generate_critical_access_report(self, wb):
        """Generate critical access report."""
        ws = wb.active
        ws.title = "Critical Access"

        self._add_report_header(ws, "Critical Transaction Access")

        # Get data
        critical = self.security_tools.check_critical_tcodes()

        # Add summary
        ws['A4'] = "Summary"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = f"TCodes Checked: {critical.get('tcodes_checked', 0)}"

        # Add data
        row = 7
        for result in critical.get("results", []):
            ws.cell(row=row, column=1, value=f"Transaction: {result.get('tcode', '')}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            ws.cell(row=row, column=1, value=f"Description: {result.get('description', '')}")
            row += 1

            risk_cell = ws.cell(row=row, column=1, value=f"Risk: {result.get('risk', '')}")
            row += 1

            ws.cell(row=row, column=1, value=f"User Count: {result.get('user_count', 0)}")
            row += 1

            # User table
            if result.get("users"):
                headers = ["Username", "Via Roles"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                self._apply_header_style(ws, row, len(headers))
                row += 1

                for user in result.get("users", [])[:10]:
                    ws.cell(row=row, column=1, value=user.get("username", ""))
                    ws.cell(row=row, column=2, value=", ".join(user.get("via_roles", [])[:3]))
                    row += 1

            row += 2  # Space between sections

        self._auto_column_width(ws)

    def _generate_full_report(self, wb):
        """Generate full security report with all findings."""
        # Generate risk summary first
        risk_summary = self.generate_risk_summary()

        # Sheet 1: Executive Summary
        ws = wb.active
        ws.title = "Executive Summary"

        self._add_report_header(ws, "Complete Security Assessment")

        # System info
        ws['A4'] = "System Information"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = risk_summary.get("system_info", "Unknown")

        # Risk overview
        ws['A7'] = "Risk Overview"
        ws['A7'].font = Font(bold=True, size=12)

        ws['A8'] = "Overall Risk Level:"
        risk_cell = ws['B8'] = risk_summary.get("overall_risk", "Unknown")
        ws['B8'].font = Font(bold=True, size=14)

        ws['A9'] = "Risk Score:"
        ws['B9'] = f"{risk_summary.get('risk_score', 0)}/10"

        summary = risk_summary.get("summary", {})
        ws['A11'] = "Critical Issues:"
        ws['B11'] = summary.get("critical_issues", 0)
        ws['A12'] = "High Issues:"
        ws['B12'] = summary.get("high_issues", 0)
        ws['A13'] = "Medium Issues:"
        ws['B13'] = summary.get("medium_issues", 0)

        # Top actions
        ws['A15'] = "Top 5 Recommended Actions"
        ws['A15'].font = Font(bold=True, size=12)

        for i, action in enumerate(risk_summary.get("top_5_actions", []), 1):
            ws[f'A{15 + i}'] = f"{i}. {action}"

        self._auto_column_width(ws)

        # Sheet 2: Detailed Findings
        ws2 = wb.create_sheet("Detailed Findings")

        headers = ["Category", "Count", "Risk", "Description", "Action Required"]
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
        self._apply_header_style(ws2, 1, len(headers))

        row = 2
        for finding in risk_summary.get("findings", []):
            ws2.cell(row=row, column=1, value=finding.get("category", ""))
            ws2.cell(row=row, column=2, value=finding.get("count", 0))
            risk_cell = ws2.cell(row=row, column=3, value=finding.get("risk", ""))
            self._apply_risk_color(ws2, risk_cell, finding.get("risk", ""))
            ws2.cell(row=row, column=4, value=finding.get("description", ""))
            ws2.cell(row=row, column=5, value=finding.get("action", ""))
            row += 1

        self._auto_column_width(ws2)

        # Sheet 3: Recommendations
        ws3 = wb.create_sheet("Recommendations")

        ws3['A1'] = "Security Recommendations"
        ws3['A1'].font = Font(bold=True, size=14)

        recommendations = [
            "1. Remove SAP_ALL and SAP_NEW profiles from all users immediately",
            "2. Lock or delete dormant user accounts inactive for 90+ days",
            "3. Implement proper segregation of duties controls",
            "4. Lock all default SAP users (SAP*, DDIC, etc.)",
            "5. Enforce strong password policies",
            "6. Review and restrict access to critical transactions",
            "7. Implement regular user access reviews",
            "8. Monitor and investigate failed background jobs",
            "9. Secure RFC destinations with proper authentication",
            "10. Maintain audit logs and review regularly"
        ]

        for i, rec in enumerate(recommendations, 3):
            ws3[f'A{i}'] = rec

        self._auto_column_width(ws3)
